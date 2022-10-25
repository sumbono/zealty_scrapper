from bs4 import BeautifulSoup
import pandas as pd
import openpyxl

def scrape_property_assesment(table_html):
    soup = BeautifulSoup(table_html, "html.parser")
    th = soup.find_all("th")
    thlist = [th_item.get_text() for th_item in th]
    tr = soup.find_all("tr")

    # List that contain format_dict(s) for pandas dataframe
    list_of_dict_table = []

    # Looping for extract data each td and create list of dict
    for index_tr_item, tr_item in enumerate(tr):
        # Find all td tag as column cell data
        td = tr_item.find_all("td")
        # Formatting dict to be appended in list_of_dict_table
        format_dict = {}
        for index_td_item, td_item in enumerate(td):
            column_data = td_item.get_text()
            format_dict[thlist[index_td_item]] = column_data
        if format_dict != {}:
            list_of_dict_table.append(format_dict)

    # Pandas Excel without merge operation
    df = pd.DataFrame(list_of_dict_table)
    df.to_excel("property_assesment.xlsx", index=False)

    # Workbook Openpyxl with merge operation
    wb = openpyxl.load_workbook("property_assesment.xlsx")
    sheet = wb.active  
    assesment_result = soup.find("div",{"style":"font-size: 12pt; text-align: center; font-weight: bold; margin-top: 8px;"}).get_text()

    sheet.merge_cells(start_row=len(df)+2, start_column=1, 
                                end_row=len(df)+2, end_column=len(df)+2)
    cell = sheet.cell(row=len(df)+2, column=1)  
    cell.value = assesment_result
    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center') 

    print("Exporting Property Assesment table to xlsx output..")
    wb.save("property_assesment.xlsx")