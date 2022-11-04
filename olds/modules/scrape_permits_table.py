import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import Workbook  
from openpyxl.styles import Alignment  

from core.config import BaseConfig

url_detail = "https://www.zealty.ca/mls-R2720295/4838-BELMONT-AVENUE-Vancouver-BC/"

def scrape_permits_table(table_html_string):
    soup = BeautifulSoup(table_html_string, "html.parser")
    tr = soup.find_all("tr")
    # List that contain format_dict(s) for pandas dataframe
    list_of_dict_table = []
    # Workbook Openpyxl
    wb = Workbook()  
    sheet = wb.active  

    # Looping for inserting data to xlsx
    for index_tr_item, tr_item in enumerate(tr):
        # Dict formatting for pandas dataframe
        format_dict = {}
        # Find td tag
        find_td = tr_item.find_all("td")
        # find td per row
        if index_tr_item % 2 == 0:
            column1 = find_td[0]
            column2 = find_td[1]

            if column1.findChild("div",{"style": "font-size: smaller;"}) :
                parent_element_col1 = str(column1).split('<div style="font-size: smaller;">')[0].replace("<div>","")
                parent_element_col1 = BeautifulSoup(parent_element_col1, "html.parser").get_text()
                child_element_col1 = column1.findChild("div",{"style": "font-size: smaller;"}).get_text()
                column1_data = parent_element_col1 + "|" + child_element_col1
                # format_dict[0] = column1_data
                cell1 = sheet.cell(row=index_tr_item+1, column=1)  
                cell1.value = column1_data
                cell1.alignment = Alignment(horizontal='center', vertical='center') 

            if column2.findChild("div",{"style": "font-size: smaller;"}) :
                parent_element_col2 = str(column2).split('<div style="font-size: smaller;">')[0].replace("<div>","")
                parent_element_col2 = BeautifulSoup(parent_element_col2, "html.parser").get_text()
                child_element_col2 = column2.findChild("div",{"style": "font-size: smaller;"}).get_text()
                column2_data = parent_element_col2 + "|" + child_element_col2
                #     format_dict[1] = column2_data
                cell2 = sheet.cell(row=index_tr_item+1, column=2)  
                cell2.value = column2_data
                cell2.alignment = Alignment(horizontal='center', vertical='center') 
            # list_of_dict_table.append(format_dict)

        else :
            sheet.merge_cells(start_row=index_tr_item+1, start_column=1, 
                                end_row=index_tr_item+1, end_column=2)
            cell = sheet.cell(row=index_tr_item+1, column=1)  
            cell.value = find_td[0].text.strip() 

    # Export xlsx
    print("Exporting Permits table to xlsx output..")
    wb.save('{}/temp/permits_table.xlsx'.format(BaseConfig.BASE_DIR))  
    wb.close()